import openpyxl

def make_xls(colnames,list):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for cols in range(1,len(colnames)+1):
        sheet.cell(1,cols,colnames[cols-1].title())
    for rows in range(2,7):
        for cols in range(1,4):
            sheet.cell(rows,cols,list[rows-1][cols-1])
            
    wb.save("MyReport.xlsx")